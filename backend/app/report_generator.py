import pandas as pd
from io import BytesIO
from datetime import datetime
import logging
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

class ReportGenerator:
    """Generate PDF and XLSX reports"""

    @classmethod
    def generate_pdf_report(cls, predictions: list, filters: dict = None) -> BytesIO:
        """
        Generate a PDF report from predictions.
        Returns: BytesIO object containing PDF data
        """
        try:
            # Create in-memory PDF
            pdf_buffer = BytesIO()
            doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
            
            # Container for PDF elements
            elements = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1f2937'),
                spaceAfter=30,
                alignment=1  # Center
            )
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#374151'),
                spaceAfter=12,
                spaceBefore=12
            )
            
            # Title
            elements.append(Paragraph("Credit Union Member Churn Prediction Report", title_style))
            elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))
            
            # Summary Statistics
            df_pred = pd.DataFrame(predictions)
            
            elements.append(Paragraph("Executive Summary", heading_style))
            summary_data = [
                ["Metric", "Value"],
                ["Total Members Analyzed", str(len(df_pred))],
                ["High Risk (>70%)", str(len(df_pred[df_pred['churn_probability'] >= 0.7]))],
                ["Medium Risk (50-70%)", str(len(df_pred[(df_pred['churn_probability'] >= 0.5) & (df_pred['churn_probability'] < 0.7)]))],
                ["Low Risk (30-50%)", str(len(df_pred[(df_pred['churn_probability'] >= 0.3) & (df_pred['churn_probability'] < 0.5)]))],
                ["Safe (<30%)", str(len(df_pred[df_pred['churn_probability'] < 0.3]))],
            ]
            
            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 0.3*inch))
            
            # Risk Distribution
            risk_counts = df_pred['risk_bucket'].value_counts().to_dict()
            elements.append(Paragraph("Risk Distribution", heading_style))
            risk_data = [["Risk Level", "Count"]]
            for risk, count in risk_counts.items():
                risk_data.append([str(risk), str(count)])
            
            risk_table = Table(risk_data, colWidths=[3*inch, 2*inch])
            risk_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(risk_table)
            elements.append(Spacer(1, 0.3*inch))
            
            # Top At-Risk Members
            elements.append(Paragraph("Top 10 At-Risk Members", heading_style))
            top_risk = df_pred.nlargest(10, 'churn_probability')[
                ['member_id', 'age', 'country', 'balance', 'churn_probability', 'risk_bucket']
            ].copy()
            top_risk['churn_probability'] = top_risk['churn_probability'].apply(lambda x: f"{x:.2%}")
            
            top_risk_data = [["Member ID", "Age", "Country", "Balance", "Churn %", "Risk Level"]]
            for _, row in top_risk.iterrows():
                top_risk_data.append([
                    str(row['member_id']),
                    str(row['age']),
                    str(row['country']),
                    f"${row['balance']:,.0f}",
                    str(row['churn_probability']),
                    str(row['risk_bucket'])
                ])
            
            top_risk_table = Table(top_risk_data, colWidths=[1*inch, 0.7*inch, 0.9*inch, 1.2*inch, 1*inch, 1.2*inch])
            top_risk_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            elements.append(top_risk_table)
            
            # Build PDF
            doc.build(elements)
            pdf_buffer.seek(0)
            
            logger.info("✅ PDF report generated successfully")
            return pdf_buffer
            
        except Exception as e:
            logger.error(f"Error generating PDF report: {e}")
            raise

    @classmethod
    def generate_xlsx_report(cls, predictions: list, filters: dict = None) -> BytesIO:
        """
        Generate an XLSX report from predictions.
        Returns: BytesIO object containing XLSX data
        """
        try:
            # Create workbook
            wb = Workbook()
            
            # Sheet 1: Summary
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            df_pred = pd.DataFrame(predictions)
            
            # Header styling
            header_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Summary section
            ws_summary['A1'] = "Credit Union Member Churn Prediction Report"
            ws_summary['A1'].font = Font(bold=True, size=14)
            ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            ws_summary['A4'] = "Executive Summary"
            ws_summary['A4'].font = Font(bold=True, size=12)
            
            summary_metrics = [
                ["Total Members Analyzed", len(df_pred)],
                ["High Risk (>70%)", len(df_pred[df_pred['churn_probability'] >= 0.7])],
                ["Medium Risk (50-70%)", len(df_pred[(df_pred['churn_probability'] >= 0.5) & (df_pred['churn_probability'] < 0.7)])],
                ["Low Risk (30-50%)", len(df_pred[(df_pred['churn_probability'] >= 0.3) & (df_pred['churn_probability'] < 0.5)])],
                ["Safe (<30%)", len(df_pred[df_pred['churn_probability'] < 0.3])],
            ]
            
            for idx, (label, value) in enumerate(summary_metrics, start=5):
                ws_summary[f'A{idx}'] = label
                ws_summary[f'B{idx}'] = value
                ws_summary[f'A{idx}'].font = Font(bold=True)
            
            # Sheet 2: All Predictions
            ws_predictions = wb.create_sheet("Predictions")
            
            # Column headers
            columns = ['member_id', 'age', 'country', 'credit_score', 'balance', 'churn_probability', 'risk_bucket', 'days_to_churn']
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws_predictions.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            
            # Data rows
            for row_idx, pred in enumerate(predictions, start=2):
                ws_predictions.cell(row=row_idx, column=1).value = str(pred.get('member_id', ''))
                ws_predictions.cell(row=row_idx, column=2).value = pred.get('age', '')
                ws_predictions.cell(row=row_idx, column=3).value = pred.get('country', '')
                ws_predictions.cell(row=row_idx, column=4).value = pred.get('credit_score', '')
                ws_predictions.cell(row=row_idx, column=5).value = pred.get('balance', '')
                
                churn_prob_cell = ws_predictions.cell(row=row_idx, column=6)
                churn_prob_cell.value = pred.get('churn_probability', 0)
                churn_prob_cell.number_format = '0.00%'
                
                ws_predictions.cell(row=row_idx, column=7).value = pred.get('risk_bucket', '')
                ws_predictions.cell(row=row_idx, column=8).value = pred.get('days_to_churn', '')
            
            # Adjust column widths
            ws_predictions.column_dimensions['A'].width = 12
            ws_predictions.column_dimensions['B'].width = 10
            ws_predictions.column_dimensions['C'].width = 12
            ws_predictions.column_dimensions['D'].width = 12
            ws_predictions.column_dimensions['E'].width = 12
            ws_predictions.column_dimensions['F'].width = 12
            ws_predictions.column_dimensions['G'].width = 15
            ws_predictions.column_dimensions['H'].width = 12
            
            # Save to BytesIO
            xlsx_buffer = BytesIO()
            wb.save(xlsx_buffer)
            xlsx_buffer.seek(0)
            
            logger.info("✅ XLSX report generated successfully")
            return xlsx_buffer
            
        except Exception as e:
            logger.error(f"Error generating XLSX report: {e}")
            raise